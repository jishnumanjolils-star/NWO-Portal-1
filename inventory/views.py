from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.db.models import Count, Sum, Avg, F, Q
from django.db import IntegrityError
from .models import NWO, TelephoneExchange, Cable, Equipment, EBCircuit, MobileBTS, JunctionBox, LIU, Fiber, Splicing, FTTH, OHMaintenanceEntry, OHMaintenanceActivity, OHMaintenanceRateMaster, UserProfile
from .forms import LIUForm, JBForm, CableForm, EquipmentForm, CircuitForm, BTSForm, Non4GBTSForm, FTTHForm, ChangePasswordForm, OHMaintenanceEntryForm, OHMaintenanceActivityFormSet
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
import logging
# import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
import json
from django.views.decorators.http import require_http_methods

class DivisionRequiredMixin:
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        queryset = super().get_queryset()
        if not self.request.user.is_superuser:
            if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
                division = self.request.user.profile.division
                # Filter based on the model type
                if self.model == Cable:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == Equipment:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == EBCircuit:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == MobileBTS:
                    queryset = queryset.filter(Q(maan_node__te__nwo=division) | Q(te__nwo=division)).distinct()
                elif self.model == JunctionBox:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == LIU:
                    queryset = queryset.filter(te__nwo=division)
                elif self.model == FTTH:
                    queryset = queryset.filter(division=division)
                elif self.model == OHMaintenanceEntry:
                    queryset = queryset.filter(division=division)
        return queryset

class TEContextMixin:
    def get_initial(self):
        initial = super().get_initial()
        te_id = self.request.GET.get('te')
        if te_id:
            initial['te'] = te_id
        return initial


def _get_division_default_password(division_name):
    defaults = {
        'NWO CENTRAL': 'Nwo@Central@2026!',
        'NWO PALARIVATTOM': 'Nwo@Palarivattom@2026!',
        'NWO KOCHI': 'Nwo@Kochi2026!',
        'NWO TRIPUNITHARA': 'Nwo@Tripunithura@2026!',
        'NWO ANGAMALY': 'Nwo@Angamaly@2026!',
        'NWO THODUPUZHA': 'Nwo@Thodupuzha@2026!',
        'NWO ALUVA': 'Nwo@Aluva@2026!',
        'NWO MOOVATTUPUZHA': 'Nwo@Moovattupuzha@2026!',
        'NWO ADIMALY': 'Nwo@Adimaly@2026!',
        'NWO KATTAPPANA': 'Nwo@Kattappana@2026!',
    }
    return defaults.get(division_name)


@login_required
@require_http_methods(["POST"])
def reset_password(request):
    profile = getattr(request.user, 'profile', None)
    if not profile or not profile.division:
        messages.error(request, 'Unable to reset password: division is not configured for your account.')
        return redirect('change_password')

    default_password = _get_division_default_password(profile.division.name)
    if not default_password:
        messages.error(request, 'Unable to reset password: default password is not configured for your division.')
        return redirect('change_password')

    request.user.set_password(default_password)
    request.user.save()

    profile.force_password_change = True
    profile.last_password_reset = timezone.now()
    profile.save(update_fields=['force_password_change', 'last_password_reset'])

    logger = logging.getLogger(__name__)
    logger.info('Password reset via division default for user=%s division=%s', request.user.username, profile.division.name)

    logout(request)
    messages.success(request, 'Password reset successfully. Please login using the default division password and change it immediately.')
    return redirect('login')


@login_required
@require_http_methods(["GET", "POST"])
def change_password(request):
    if request.method == "GET":
        form = ChangePasswordForm(user=request.user)
        force_change = bool(getattr(request.user, 'profile', None) and request.user.profile.force_password_change)
        return render(request, "inventory/change_password.html", {"form": form, "force_password_change": force_change})

    is_json = "application/json" in (request.content_type or "")
    if is_json:
        try:
            payload = json.loads((request.body or b"{}").decode("utf-8"))
        except Exception:
            return HttpResponseBadRequest("Invalid JSON")
        if "new_password" in payload and "confirm_new_password" not in payload:
            payload["confirm_new_password"] = payload.get("new_password")
        form = ChangePasswordForm(payload, user=request.user)
    else:
        form = ChangePasswordForm(request.POST, user=request.user)

    if form.is_valid():
        request.user.set_password(form.cleaned_data["new_password"])
        request.user.save()
        profile = getattr(request.user, 'profile', None)
        if profile and profile.force_password_change:
            profile.force_password_change = False
            profile.save(update_fields=['force_password_change'])
        logout(request)
        msg = "Password changed successfully. Please login again."
        if is_json:
            return JsonResponse({"detail": msg})
        messages.success(request, msg)
        return redirect("login")

    if is_json:
        return JsonResponse({"errors": form.errors}, status=400)
    force_change = bool(getattr(request.user, 'profile', None) and request.user.profile.force_password_change)
    return render(request, "inventory/change_password.html", {"form": form, "force_password_change": force_change})


@login_required
def te_jb_routes(request, te_id):
    te = get_object_or_404(TelephoneExchange, id=te_id)
    if not request.user.is_superuser and hasattr(request.user, "profile") and getattr(request.user.profile, "division_id", None):
        if te.nwo_id != request.user.profile.division_id:
            return HttpResponse(status=403)

    jbs = list(
        JunctionBox.objects.filter(te=te).prefetch_related("input_cables", "output_cables")
    )

    def build_routes(mode):
        nodes = [jb for jb in jbs if jb.cable_mode == mode]
        if not nodes:
            return []

        inputs = {jb.id: set(c.id for c in jb.input_cables.all()) for jb in nodes}
        outputs = {jb.id: set(c.id for c in jb.output_cables.all()) for jb in nodes}
        edges = {jb.id: [] for jb in nodes}
        incoming = {jb.id: 0 for jb in nodes}

        for a in nodes:
            out_set = outputs.get(a.id, set())
            if not out_set:
                continue
            for b in nodes:
                if a.id == b.id:
                    continue
                if out_set & inputs.get(b.id, set()):
                    edges[a.id].append(b.id)
                    incoming[b.id] += 1

        starts = [jb_id for jb_id, deg in incoming.items() if deg == 0]
        if not starts:
            starts = [jb.id for jb in nodes]

        id_to = {jb.id: jb for jb in nodes}
        routes = []

        def dfs(cur, path, seen):
            if len(routes) >= 100:
                return
            nxts = edges.get(cur, [])
            if not nxts:
                routes.append([id_to[i] for i in path])
                return
            extended = False
            for nxt in nxts:
                if nxt in seen:
                    continue
                extended = True
                dfs(nxt, path + [nxt], seen | {nxt})
            if not extended:
                routes.append([id_to[i] for i in path])

        for s in starts:
            dfs(s, [s], {s})

        uniq = []
        seen_keys = set()
        for r in routes:
            key = tuple(j.id for j in r)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            uniq.append(r)
        return uniq

    ug_routes = build_routes("UG")
    oh_routes = build_routes("OH")

    def points(mode):
        pts = []
        for jb in jbs:
            if jb.cable_mode != mode:
                continue
            if jb.latitude is None or jb.longitude is None:
                continue
            pts.append(
                {
                    "id": jb.id,
                    "jb_id": jb.jb_id,
                    "name": jb.jb_name or "",
                    "lat": float(jb.latitude),
                    "lng": float(jb.longitude),
                    "url": reverse("jb_detail", args=[jb.id]),
                }
            )
        return pts

    return render(
        request,
        "inventory/te_jb_routes.html",
        {
            "te": te,
            "ug_routes": ug_routes,
            "oh_routes": oh_routes,
            "ug_points_json": json.dumps(points("UG")),
            "oh_points_json": json.dumps(points("OH")),
        },
    )

@login_required
def dashboard(request):
    # 1. Identify User Division
    user_division = None
    if hasattr(request.user, 'profile') and request.user.profile.division:
        user_division = request.user.profile.division

    # Auto-assign/match unlinked MobileBTS sites on dashboard load
    if user_division:
        _auto_assign_bts_helper(user_division)

    # 2. Fetch Divisions
    if user_division:
        nwos_to_process = NWO.objects.filter(id=user_division.id)
    else:
        nwos_to_process = NWO.objects.all()

    # 3. Sort according to preference
    nwo_order = [
        'NWO CENTRAL', 'NWO PALARIVATTOM', 'NWO KOCHI', 'NWO TRIPUNITHARA',
        'NWO ANGAMALY', 'NWO THODUPUZHA', 'NWO ALUVA', 'NWO MOOVATTUPUZHA',
        'NWO ADIMALY', 'NWO KATTAPPANA'
    ]
    
    nwo_map = {n.name: n for n in nwos_to_process}
    sorted_nwos = []
    for code in nwo_order:
        if code in nwo_map:
            sorted_nwos.append(nwo_map[code])
    
    # Add any other divisions that might not be in the order list
    for nwo in nwos_to_process:
        if nwo not in sorted_nwos:
            sorted_nwos.append(nwo)

    # 4. Build data structure
    nwo_data_list = []
    for nwo in sorted_nwos:
        # Use explicit queryset to avoid any related_name/attribute resolution issues
        exchanges = TelephoneExchange.objects.filter(nwo=nwo).order_by('name')
        te_list = []
        for te in exchanges:
            te_list.append({
                'id': te.id,
                'name': te.name,
                'cable_count': te.cables.count(),
                'equipment_count': te.equipments.count(),
                'circuit_count': EBCircuit.objects.filter(te=te).count(),
                'bts_count': MobileBTS.objects.filter(Q(maan_node__te=te) | Q(te=te)).distinct().count(),
                'ftth_count': FTTH.objects.filter(te=te).count(),
            })
        
        nwo_data_list.append({
            'display_name': nwo.name,
            'te_data': te_list
        })

    # 5. Global Stats
    if user_division:
        total_tes = TelephoneExchange.objects.filter(nwo=user_division).count()
        total_circuits = EBCircuit.objects.filter(te__nwo=user_division).count()
        total_cables = Cable.objects.filter(te__nwo=user_division).count()
        total_equipment = Equipment.objects.filter(te__nwo=user_division).count()
        # For BTS, we count linked ones. For FTTH, we use the direct division relation.
        total_bts = MobileBTS.objects.filter(Q(maan_node__te__nwo=user_division) | Q(te__nwo=user_division)).distinct().count()
        total_ftth = FTTH.objects.filter(division=user_division).count()
    else:
        total_tes = TelephoneExchange.objects.count()
        total_circuits = EBCircuit.objects.count()
        total_cables = Cable.objects.count()
        total_equipment = Equipment.objects.count()
        total_bts = MobileBTS.objects.count()
        total_ftth = FTTH.objects.count()

    # System Diagnostic data
    all_bts_debug = []
    if request.user.is_superuser:
        bts_qs = MobileBTS.objects.select_related('te')
        if user_division:
            bts_qs = bts_qs.filter(Q(maan_node__te__nwo=user_division) | Q(te__nwo=user_division)).distinct()
        all_bts_debug = bts_qs.order_by('te__name', 'rp_id')

    context = {
        'nwo_data_list': nwo_data_list,
        'total_tes': total_tes,
        'total_circuits': total_circuits,
        'total_cables': total_cables,
        'total_equipment': total_equipment,
        'total_bts': total_bts,
        'total_ftth': total_ftth,
        'is_filtered': user_division is not None,
        'user_division_name': user_division.name if user_division else "All Divisions",
        'all_bts_debug': all_bts_debug,
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required
def te_dashboard(request, te_id):
    te = get_object_or_404(TelephoneExchange, id=te_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this TE.")
                return redirect('dashboard')
    
    cables = Cable.objects.filter(te=te)
    equipments = Equipment.objects.filter(te=te)
    bts_count = MobileBTS.objects.filter(Q(maan_node__te=te) | Q(te=te)).distinct().count()
    
    # Cable summary
    ug_summary = cables.filter(mode='UG').values('cable_type').annotate(count=Count('id')).order_by('cable_type')
    oh_summary = cables.filter(mode='OH').values('cable_type').annotate(count=Count('id')).order_by('cable_type')
    
    ug_dict = {item['cable_type']: item['count'] for item in ug_summary}
    oh_dict = {item['cable_type']: item['count'] for item in oh_summary}
    
    # Cable categories
    in_cables = cables.filter(category='IN').count()
    out_cables = cables.filter(category='OUT').count()
    tie_cables = cables.filter(category='TIE').count()
    
    # Convert Decimals to floats for JS compatibility in JBs
    jbs_raw = JunctionBox.objects.filter(te=te)
    jbs = []
    jbs_list = []
    for jb in jbs_raw:
        jb_item = {
            'jb_id': jb.jb_id,
            'latitude': float(jb.latitude) if jb.latitude else None,
            'longitude': float(jb.longitude) if jb.longitude else None,
        }
        jbs.append(jb_item)
        if jb_item['latitude'] and jb_item['longitude']:
            jbs_list.append(jb_item)
            
    equipments_list = []
    for eq in equipments:
        if eq.latitude and eq.longitude:
            equipments_list.append({
                'name': eq.name,
                'latitude': float(eq.latitude),
                'longitude': float(eq.longitude),
                'used_ports': eq.used_ports,
                'total_ports': eq.total_ports,
            })
            
    bts_list = MobileBTS.objects.filter(Q(maan_node__te=te) | Q(te=te)).distinct()
    bts_list_json = []
    for b in bts_list:
        if b.latitude and b.longitude:
            bts_list_json.append({
                'bts_name': b.bts_name,
                'rp_id': b.rp_id,
                'site_type': b.get_site_type_display(),
                'latitude': float(b.latitude),
                'longitude': float(b.longitude),
                'place_name': b.place_name or '',
            })
            
    eb_circuits_json = []
    for eb in EBCircuit.objects.filter(te=te):
        if eb.latitude and eb.longitude:
            eb_circuits_json.append({
                'client_name': eb.client_name or 'Unknown',
                'circuit_type': eb.circuit_type or 'Unknown',
                'bandwidth': eb.bandwidth or '',
                'latitude': float(eb.latitude),
                'longitude': float(eb.longitude),
                'customer_premise_location': eb.customer_premise_location or '',
            })
    
    # Fetch route points for cables in this TE
    cable_routes = []
    cable_routes_list = []
    for cable in cables:
        # Convert Decimals to floats for JS compatibility
        points = [[float(p[0]), float(p[1])] for p in cable.route_points.all().values_list('latitude', 'longitude', flat=False)]
        if points:
            route_item = {
                'name': cable.name,
                'type': cable.cable_type,
                'mode': cable.mode,
                'category': cable.category or '',
                'points': points
            }
            cable_routes.append(route_item)
            cable_routes_list.append(route_item)
            
    context = {
        'te': te,
        'cables': cables,
        'equipments': equipments,
        'bts_count': bts_count,
        'bts_list': bts_list,
        'ug_summary': ug_dict,
        'oh_summary': oh_dict,
        'in_cables': in_cables,
        'out_cables': out_cables,
        'tie_cables': tie_cables,
        'jbs': jbs,
        'cable_routes': cable_routes,
        'jbs_json': json.dumps(jbs_list),
        'equipments_json': json.dumps(equipments_list),
        'cable_routes_json': json.dumps(cable_routes_list),
        'bts_assets_json': json.dumps(bts_list_json),
        'eb_circuits_json': json.dumps(eb_circuits_json),
    }
    return render(request, 'inventory/te_dashboard.html', context)

class CableListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = Cable
    template_name = 'inventory/cable_list.html'
    context_object_name = 'cables'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('te', 'te__nwo')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        # Cable type filter
        cable_type = self.request.GET.get('cable_type')
        if cable_type:
            queryset = queryset.filter(cable_type=cable_type)
        
        # Mode filter (OH/UG)
        mode = self.request.GET.get('mode')
        if mode:
            queryset = queryset.filter(mode=mode)
        
        # Category filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_term'] = self.request.GET.get('search', '')
        return context

class EquipmentListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = Equipment
    template_name = 'inventory/equipment_list.html'
    context_object_name = 'equipments'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('te', 'te__nwo')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        # Equipment type filter
        eq_type = self.request.GET.get('equipment_type')
        if eq_type:
            queryset = queryset.filter(equipment_type=eq_type)
        
        # TE filter
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
        
        return queryset

class CircuitListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = EBCircuit
    template_name = 'inventory/circuit_list.html'
    context_object_name = 'circuits'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('te', 'te__nwo', 'cable', 'equipment')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(client_name__icontains=search) | Q(lc_id__icontains=search)
            )
        
        # Circuit type filter
        circuit_type = self.request.GET.get('circuit_type')
        if circuit_type:
            queryset = queryset.filter(circuit_type=circuit_type)
        
        # TE filter
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Base queryset of circuits for counting, respecting division
        base_qs = EBCircuit.objects.all()
        if not self.request.user.is_superuser:
            if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
                division = self.request.user.profile.division
                base_qs = base_qs.filter(te__nwo=division)
                
        # Group by and count
        from django.db.models import Count
        counts = base_qs.values('circuit_type').annotate(count=Count('id'))
        counts_dict = {item['circuit_type']: item['count'] for item in counts}
        
        categories = [
            {'type': 'INTERNET LC', 'badge_class': 'bg-tint-blue'},
            {'type': 'P2P LC', 'badge_class': 'bg-tint-cyan'},
            {'type': 'P2P LC ACROSS STATE', 'badge_class': 'bg-tint-violet'},
            {'type': 'MPLS VPN', 'badge_class': 'bg-tint-emerald'},
            {'type': 'ISDN PRI', 'badge_class': 'bg-tint-amber'},
        ]
        
        dashboard_counts = []
        for cat in categories:
            dashboard_counts.append({
                'type': cat['type'],
                'count': counts_dict.get(cat['type'], 0),
                'badge_class': cat['badge_class']
            })
            
        context['dashboard_counts'] = dashboard_counts
        return context


class BTSListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = MobileBTS
    template_name = 'inventory/bts_list.html'
    context_object_name = 'bts_list'

    def get_queryset(self):
        if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
            _auto_assign_bts_helper(self.request.user.profile.division)
        queryset = super().get_queryset().filter(site_type='4G').select_related('maan_node', 'maan_node__te')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(rp_id__icontains=search) | 
                Q(bts_name__icontains=search)
            )
        
        # Ring filter
        is_ring = self.request.GET.get('is_ring')
        if is_ring:
            queryset = queryset.filter(is_ring=is_ring == 'true')
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_count'] = queryset.count()
        context['ring_count'] = queryset.filter(is_ring=True).count()
        context['cef_count'] = queryset.filter(has_cef_12t=True).count()
        context['avg_power'] = queryset.aggregate(Avg('receive_power_db'))['receive_power_db__avg']
        return context

@login_required
def te_liu_setup(request, te_id):
    te = get_object_or_404(TelephoneExchange, id=te_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this TE.")
                return redirect('dashboard')

    cables = Cable.objects.filter(te=te)
    
    if request.method == 'POST':
        liu_count = int(request.POST.get('liu_count', 0))
        for i in range(1, liu_count + 1):
            name = request.POST.get(f'liu_{i}_name')
            cable_id = request.POST.get(f'liu_{i}_cable')
            cable_manual = request.POST.get(f'liu_{i}_cable_manual')
            capacity = int(request.POST.get(f'liu_{i}_capacity', 0))
            remarks = request.POST.get(f'liu_{i}_remarks')
            
            if name and capacity and (cable_id or cable_manual):
                cable = None
                cable_manual_entry = None
                
                if cable_id:
                    # Use selected cable from dropdown
                    cable = get_object_or_404(Cable, id=cable_id)
                else:
                    # Use manual cable entry
                    cable_manual_entry = cable_manual
                
                liu = LIU.objects.create(
                    te=te,
                    name=name,
                    cable=cable,
                    cable_manual_entry=cable_manual_entry,
                    capacity=capacity,
                    remarks=remarks or f"Setup for {te.name}"
                )
                
                # Update ports with fiber/circuit data
                for port_num in range(1, capacity + 1):
                    circuit = request.POST.get(f'liu_{i}_fiber_{port_num}_circuit')
                    sys_port = request.POST.get(f'liu_{i}_fiber_{port_num}_port')
                    otdr_distance = request.POST.get(f'liu_{i}_fiber_{port_num}_otdr_distance')
                    otdr_image = request.FILES.get(f'liu_{i}_fiber_{port_num}_otdr_image')
                    
                    if circuit or sys_port or otdr_distance or otdr_image:
                        port = liu.ports.get(port_number=port_num)
                        port.connected_to = sys_port
                        port.remarks = circuit
                        port.otdr_distance = otdr_distance
                        if otdr_image:
                            port.otdr_image = otdr_image
                        if circuit or sys_port:
                            port.status = 'Used'
                        port.save()
        
        messages.success(request, f"Successfully set up {liu_count} LIUs for {te.name}. Please configure fiber details for each LIU below.")
        return redirect(f"{reverse_lazy('liu_list')}?te={te.id}")

    return render(request, 'inventory/te_liu_setup.html', {
        'te': te,
        'cables': cables,
        'range_10': range(1, 11)
    })


@login_required
def get_cable_details(request, cable_id):
    try:
        cable = Cable.objects.get(id=cable_id)
        if not request.user.is_superuser:
            if hasattr(request.user, 'profile') and request.user.profile.division:
                if cable.te.nwo != request.user.profile.division:
                    return JsonResponse({'error': 'Access denied'}, status=403)
        
        return JsonResponse({
            'id': cable.id,
            'name': cable.name,
            'cable_type': cable.cable_type,
            'fiber_count': cable.fiber_count,
            'structure_type': cable.structure_type,
            'mode': cable.get_mode_display(),
            'category': cable.get_category_display(),
            'te': cable.te.name,
            'connected_te': cable.connected_te.name if cable.connected_te else 'N/A',
            'otdr_distance': cable.otdr_distance_formatted or 'Not recorded',
            'otdr_image_url': cable.otdr_image.url if cable.otdr_image else None,
            'remarks': cable.remarks,
        })
    except Cable.DoesNotExist:
        return JsonResponse({'error': 'Cable not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


class LIUListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = LIU
    template_name = 'inventory/liu_list.html'
    context_object_name = 'liu_list'

    def get_queryset(self):
        queryset = super().get_queryset()
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        te_id = self.request.GET.get('te')
        if te_id:
            context['filter_te'] = get_object_or_404(TelephoneExchange, id=te_id)
        return context

class LIUCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = LIU
    form_class = LIUForm
    template_name = 'inventory/liu_form.html'
    success_url = reverse_lazy('liu_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        liu = self.object
        capacity = liu.capacity
        
        # Update ports with fiber/circuit data from the form
        for port_num in range(1, capacity + 1):
            circuit = self.request.POST.get(f'fiber_{port_num}_circuit')
            sys_port = self.request.POST.get(f'fiber_{port_num}_port')
            
            if circuit or sys_port:
                port = liu.ports.get(port_number=port_num)
                port.connected_to = sys_port
                port.remarks = circuit
                port.status = 'Used'
                port.save()
        
        messages.success(self.request, f"LIU {liu.name} created successfully with fiber details.")
        return response

@login_required
def liu_detail(request, liu_id):
    liu = get_object_or_404(LIU, id=liu_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if liu.te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this LIU.")
                return redirect('dashboard')

    ports = liu.ports.all()
    fibers = liu.cable.fibers.all() if liu.cable else Fiber.objects.none()

    if request.method == 'POST':
        for port in ports:
            fiber_id = request.POST.get(f'port_{port.id}_fiber')
            connected_to = request.POST.get(f'port_{port.id}_connected')
            status = request.POST.get(f'port_{port.id}_status')
            remarks = request.POST.get(f'port_{port.id}_remarks')

            if fiber_id:
                try:
                    port.fiber = Fiber.objects.get(id=fiber_id)
                except Fiber.DoesNotExist:
                    port.fiber = None
            else:
                port.fiber = None
            
            port.connected_to = connected_to
            port.status = status
            port.remarks = remarks
            port.save()
        
        messages.success(request, "LIU Ports updated successfully!")
        return redirect('liu_detail', liu_id=liu.id)

    return render(request, 'inventory/liu_detail.html', {
        'liu': liu,
        'ports': ports,
        'fibers': fibers
    })

class LIUUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = LIU
    form_class = LIUForm
    template_name = 'inventory/liu_form.html'
    
    def get_success_url(self):
        return reverse_lazy('liu_detail', kwargs={'liu_id': self.object.id})

    def form_valid(self, form):
        messages.success(self.request, f"LIU {self.object.name} updated successfully.")
        return super().form_valid(form)

class JBListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = JunctionBox
    template_name = 'inventory/jb_list.html'
    context_object_name = 'jbs'

class JBCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = JunctionBox
    form_class = JBForm
    template_name = 'inventory/jb_form.html'
    success_url = reverse_lazy('jb_list')

class JBUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = JunctionBox
    form_class = JBForm
    template_name = 'inventory/jb_form.html'
    success_url = reverse_lazy('jb_list')

class JBDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = JunctionBox
    template_name = 'inventory/jb_confirm_delete.html'
    success_url = reverse_lazy('jb_list')

class JBDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = JunctionBox
    template_name = 'inventory/jb_detail.html'
    context_object_name = 'jb'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['splices'] = self.object.splices.all()
        return context

@login_required
def jb_splicing(request, jb_id):
    jb = get_object_or_404(JunctionBox, id=jb_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if jb.te.nwo != request.user.profile.division:
                messages.error(request, "You do not have permission to access this Junction Box.")
                return redirect('dashboard')

    splices = jb.splices.all()
    cables = Cable.objects.filter(te=jb.te)
    
    if request.method == 'POST':
        fiber_in_id = request.POST.get('fiber_in')
        fiber_out_id = request.POST.get('fiber_out')
        remarks = request.POST.get('remarks')
        
        if fiber_in_id and fiber_out_id:
            fiber_in = get_object_or_404(Fiber, id=fiber_in_id)
            fiber_out = get_object_or_404(Fiber, id=fiber_out_id)
            
            Splicing.objects.create(
                jb=jb,
                fiber_in=fiber_in,
                fiber_out=fiber_out,
                remarks=remarks
            )
            messages.success(request, "Splice recorded successfully!")
            return redirect('jb_splicing', jb_id=jb.id)
            
    return render(request, 'inventory/jb_splicing.html', {
        'jb': jb,
        'splices': splices,
        'cables': cables
    })

@login_required
def api_get_fibers(request, cable_id):
    cable = get_object_or_404(Cable, id=cable_id)
    
    # Division check
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            if cable.te.nwo != request.user.profile.division:
                return JsonResponse({'error': 'Permission denied'}, status=403)

    fibers = cable.fibers.all().values('id', 'fiber_number', 'status', 'is_used')
    return JsonResponse({'fibers': list(fibers)})

class BTSCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = MobileBTS
    form_class = BTSForm
    template_name = 'inventory/bts_form.html'
    success_url = reverse_lazy('bts_list')

class BTSUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = MobileBTS
    form_class = BTSForm
    template_name = 'inventory/bts_form.html'
    success_url = reverse_lazy('bts_list')

class BTSDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = MobileBTS
    template_name = 'inventory/bts_confirm_delete.html'
    success_url = reverse_lazy('bts_list')

# FTTH Views
class FTTHListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = FTTH
    template_name = 'inventory/ftth_list.html'
    context_object_name = 'ftth_list'

    def get_queryset(self):
        queryset = super().get_queryset().select_related('division', 'te')
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(customer_name__icontains=search) |
                Q(landline_number__icontains=search) |
                Q(olt_name__icontains=search)
            )
        return queryset

class FTTHDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = FTTH
    template_name = 'inventory/ftth_detail.html'
    context_object_name = 'ftth'

class FTTHCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = FTTH
    form_class = FTTHForm
    template_name = 'inventory/ftth_form.html'
    success_url = reverse_lazy('ftth_list')

    def get_initial(self):
        initial = super().get_initial()
        if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
            initial['division'] = self.request.user.profile.division
        
        te_id = self.request.GET.get('te')
        if te_id:
            initial['te'] = te_id
        return initial

class FTTHUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = FTTH
    form_class = FTTHForm
    template_name = 'inventory/ftth_form.html'
    success_url = reverse_lazy('ftth_list')

class FTTHDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = FTTH
    template_name = 'inventory/ftth_confirm_delete.html'
    success_url = reverse_lazy('ftth_list')

class BTSDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = MobileBTS
    template_name = 'inventory/bts_detail.html'
    context_object_name = 'bts'


class No4GBTSListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = MobileBTS
    template_name = 'inventory/bts_no_4g_list.html'
    context_object_name = 'bts_list'

    def get_queryset(self):
        if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
            _auto_assign_bts_helper(self.request.user.profile.division)
        queryset = super().get_queryset().filter(site_type='NON_4G').select_related('te', 'te__nwo')
        
        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(rp_id__icontains=search) | 
                Q(bts_name__icontains=search)
            )
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()
        context['total_count'] = queryset.count()
        context['fiber_count'] = queryset.filter(backhaul_media='FIBER').count()
        context['microwave_count'] = queryset.filter(backhaul_media='MICROWAVE').count()
        return context


class No4GBTSDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = MobileBTS
    template_name = 'inventory/bts_no_4g_detail.html'
    context_object_name = 'bts'


class No4GBTSCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = MobileBTS
    form_class = Non4GBTSForm
    template_name = 'inventory/bts_no_4g_form.html'
    success_url = reverse_lazy('no_4g_bts_list')


class No4GBTSUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = MobileBTS
    form_class = Non4GBTSForm
    template_name = 'inventory/bts_no_4g_form.html'
    success_url = reverse_lazy('no_4g_bts_list')


class No4GBTSDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = MobileBTS
    template_name = 'inventory/bts_no_4g_confirm_delete.html'
    success_url = reverse_lazy('no_4g_bts_list')

import re

class CableFiberUpdateMixin:
    def form_valid(self, form):
        response = super().form_valid(form)
        cable = self.object
        for i in range(1, cable.fiber_count + 1):
            sys_end = self.request.POST.get(f'fiber_{i}_system_end')
            circ_name = self.request.POST.get(f'fiber_{i}_circuit_name')
            dist_str = self.request.POST.get(f'fiber_{i}_otdr_distance')
            img = self.request.FILES.get(f'fiber_{i}_otdr_image')
            
            try:
                fiber = cable.fibers.get(fiber_number=i)
            except Fiber.DoesNotExist:
                continue
                
            updated = False
            if sys_end is not None:
                if fiber.system_end != sys_end:
                    fiber.system_end = sys_end
                    updated = True
            if circ_name is not None:
                if fiber.circuit_name != circ_name:
                    fiber.circuit_name = circ_name
                    updated = True
            if dist_str is not None:
                val = None
                if dist_str.strip():
                    match = re.search(r'([\d.]+)', dist_str)
                    if match:
                        try:
                            val = float(match.group(1))
                            if 'km' in dist_str.lower():
                                val = val * 1000
                        except ValueError:
                            pass
                if fiber.otdr_distance != val:
                    fiber.otdr_distance = val
                    updated = True
            if img:
                fiber.otdr_image = img
                updated = True
                
            is_currently_used = bool((sys_end and sys_end.strip()) or (circ_name and circ_name.strip()))
            if fiber.is_used != is_currently_used:
                fiber.is_used = is_currently_used
                fiber.status = 'Used' if is_currently_used else 'Available'
                updated = True
            elif fiber.status != ('Used' if is_currently_used else 'Available'):
                fiber.status = 'Used' if is_currently_used else 'Available'
                updated = True
                
            if updated:
                fiber.save()
        return response

class CableCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CableFiberUpdateMixin, CreateView):
    model = Cable
    form_class = CableForm
    template_name = 'inventory/cable_form.html'
    success_url = reverse_lazy('cable_list')

class CableUpdateView(LoginRequiredMixin, DivisionRequiredMixin, CableFiberUpdateMixin, UpdateView):
    model = Cable
    form_class = CableForm
    template_name = 'inventory/cable_form.html'
    success_url = reverse_lazy('cable_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fibers'] = self.object.fibers.all()
        return context

class CableDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = Cable
    template_name = 'inventory/cable_confirm_delete.html'
    success_url = reverse_lazy('cable_list')

class CableDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = Cable
    template_name = 'inventory/cable_detail.html'
    context_object_name = 'cable'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fibers'] = self.object.fibers.all()
        return context

class EquipmentCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'inventory/equipment_form.html'
    success_url = reverse_lazy('equipment_list')

    def get_initial(self):
        initial = super().get_initial()
        eq_type = self.request.GET.get('type')
        if eq_type:
            initial['equipment_type'] = eq_type
        return initial

class EquipmentUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'inventory/equipment_form.html'
    success_url = reverse_lazy('equipment_list')

class EquipmentDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = Equipment
    template_name = 'inventory/equipment_confirm_delete.html'
    success_url = reverse_lazy('equipment_list')

class EquipmentDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = Equipment
    template_name = 'inventory/equipment_detail.html'
    context_object_name = 'equipment'

class CircuitCreateView(LoginRequiredMixin, DivisionRequiredMixin, TEContextMixin, CreateView):
    model = EBCircuit
    form_class = CircuitForm
    template_name = 'inventory/circuit_form.html'
    success_url = reverse_lazy('circuit_list')

class CircuitUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = EBCircuit
    form_class = CircuitForm
    template_name = 'inventory/circuit_form.html'
    success_url = reverse_lazy('circuit_list')

class CircuitDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = EBCircuit
    template_name = 'inventory/circuit_confirm_delete.html'
    success_url = reverse_lazy('circuit_list')

class CircuitDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = EBCircuit
    template_name = 'inventory/circuit_detail.html'
    context_object_name = 'circuit'

@login_required
def clear_circuits(request):
    if request.method == 'POST':
        queryset = EBCircuit.objects.all()
        if not request.user.is_superuser:
            if hasattr(request.user, 'profile') and request.user.profile.division:
                division = request.user.profile.division
                queryset = queryset.filter(te__nwo=division)
            else:
                messages.error(request, "You do not have permission to delete circuits.")
                return redirect('circuit_list')
                
        count = queryset.count()
        queryset.delete()
        messages.success(request, f"Successfully deleted {count} EB circuits.")
        return redirect('circuit_list')
        
    messages.error(request, "Invalid request method.")
    return redirect('circuit_list')

@login_required
def analytics(request):
    from django.db.models import F, ExpressionWrapper, DurationField, Avg, Count, Sum, Q
    from django.utils import timezone
    
    division = None
    if request.user.is_superuser:
        division_id = request.GET.get('division')
        if division_id:
            division = NWO.objects.filter(id=division_id).first()
    else:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division

    # Fault Tracking Analytics
    fault_entries = OHMaintenanceEntry.objects.filter(fault_occurrence_datetime__isnull=False)
    if division:
        fault_entries = fault_entries.filter(division=division)
    
    # 1. Fault Duration Analysis (Closed Faults)
    closed_faults = fault_entries.filter(fault_clearance_datetime__isnull=False).annotate(
        duration=ExpressionWrapper(
            F('fault_clearance_datetime') - F('fault_occurrence_datetime'),
            output_field=DurationField()
        )
    )
    
    # 2. Division-wise MTTR
    division_mttr = closed_faults.values('division__name').annotate(
        avg_mttr=Avg('duration'),
        total_closed=Count('id')
    ).order_by('avg_mttr')

    # 3. Team-wise Performance
    team_performance = closed_faults.values('team_name', 'division__name').annotate(
        avg_mttr=Avg('duration'),
        total_closed=Count('id')
    ).order_by('avg_mttr')

    # 4. Repeated Fault Locations
    repeated_locations = fault_entries.values('location', 'division__name').annotate(
        fault_count=Count('id')
    ).filter(fault_count__gt=1).order_by('-fault_count')

    # 5. Fault Type Trends
    fault_type_trends = fault_entries.values('fault_type').annotate(
        count=Count('id')
    ).order_by('-count')

    # 6. Severity Distribution
    severity_dist = fault_entries.values('fault_severity').annotate(
        count=Count('id')
    ).order_by('-count')

    # Existing Analytics...
    # Mobile Tower RP ID Summary
    bts_list = MobileBTS.objects.select_related('maan_node', 'maan_node__te').all()
    if division:
        bts_list = bts_list.filter(maan_node__te__nwo=division)
    
    # Cable Summary - OH and UG separately
    oh_cables = Cable.objects.filter(mode='OH').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')
    
    ug_cables = Cable.objects.filter(mode='UG').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')

    if division:
        oh_cables = oh_cables.filter(te__nwo=division)
        ug_cables = ug_cables.filter(te__nwo=division)
    
    # Overall cable summary
    cable_queryset = Cable.objects.all()
    if division:
        cable_queryset = cable_queryset.filter(te__nwo=division)
    cable_summary = cable_queryset.values('cable_type', 'mode').annotate(total=Count('id')).order_by('cable_type')
    
    # Equipment Summary - CPAN and MAAN separately
    cpan_nodes = Equipment.objects.filter(equipment_type__startswith='CPAN').values('name', 'te__name', 'total_ports')
    maan_nodes = Equipment.objects.filter(equipment_type__startswith='MAAN').values('name', 'equipment_type', 'te__name', 'total_ports')
    
    if division:
        cpan_nodes = cpan_nodes.filter(te__nwo=division)
        maan_nodes = maan_nodes.filter(te__nwo=division)
    
    equipment_queryset = Equipment.objects.all()
    if division:
        equipment_queryset = equipment_queryset.filter(te__nwo=division)
    equipment_summary = equipment_queryset.values('equipment_type').annotate(total=Count('id')).order_by('equipment_type')
    
    # Circuit Summary by Type
    circuit_queryset = EBCircuit.objects.all()
    if division:
        circuit_queryset = circuit_queryset.filter(te__nwo=division)
    circuit_type_summary = circuit_queryset.values('circuit_type').annotate(total=Count('id')).order_by('circuit_type')
    
    # JB Summary - OH and UG separately
    jb_queryset = JunctionBox.objects.all()
    if division:
        jb_queryset = jb_queryset.filter(te__nwo=division)
    oh_jbs = jb_queryset.filter(jb_type='OH').count()
    ug_jbs = jb_queryset.filter(jb_type='UG').count()

    if request.user.is_superuser:
        divisions = list(NWO.objects.all().order_by('name'))
    else:
        divisions = [division] if division else []

    return render(request, 'inventory/analytics.html', {
        'divisions': divisions,
        'selected_division': division,
        'bts_list': bts_list,
        'cable_summary': cable_summary,
        'oh_cables': oh_cables,
        'ug_cables': ug_cables,
        'equipment_summary': equipment_summary,
        'cpan_nodes': cpan_nodes,
        'maan_nodes': maan_nodes,
        'cpan_count': cpan_nodes.count(),
        'maan_count': maan_nodes.count(),
        'circuit_type_summary': circuit_type_summary,
        'oh_jbs': oh_jbs,
        'ug_jbs': ug_jbs,
        # New Fault Tracking Context
        'division_mttr': division_mttr,
        'team_performance': team_performance,
        'repeated_locations': repeated_locations,
        'fault_type_trends': fault_type_trends,
        'severity_dist': severity_dist,
    })

@login_required
def export_analytics(request):
    division = None
    if request.user.is_superuser:
        division_id = request.GET.get('division')
        if division_id:
            division = NWO.objects.filter(id=division_id).first()
    else:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division

    bts_list = MobileBTS.objects.select_related('maan_node', 'maan_node__te').all()
    if division:
        bts_list = bts_list.filter(maan_node__te__nwo=division)

    oh_cables = Cable.objects.filter(mode='OH').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')

    ug_cables = Cable.objects.filter(mode='UG').annotate(
        total_fibers=Count('fibers'),
        used_fibers=Count('fibers', filter=Q(fibers__is_used=True)),
        free_fibers=Count('fibers', filter=Q(fibers__is_used=False))
    ).select_related('te')

    cable_queryset = Cable.objects.all()
    equipment_queryset = Equipment.objects.all()
    circuit_queryset = EBCircuit.objects.all()

    if division:
        oh_cables = oh_cables.filter(te__nwo=division)
        ug_cables = ug_cables.filter(te__nwo=division)
        cable_queryset = cable_queryset.filter(te__nwo=division)
        equipment_queryset = equipment_queryset.filter(te__nwo=division)
        circuit_queryset = circuit_queryset.filter(te__nwo=division)

    cable_summary = cable_queryset.values('cable_type', 'mode').annotate(total=Count('id')).order_by('cable_type', 'mode')
    equipment_summary = equipment_queryset.values('equipment_type').annotate(total=Count('id')).order_by('equipment_type')
    circuit_type_summary = circuit_queryset.values('circuit_type').annotate(total=Count('id')).order_by('circuit_type')

    cpan_nodes = Equipment.objects.filter(equipment_type__startswith='CPAN')
    maan_nodes = Equipment.objects.filter(equipment_type__startswith='MAAN')
    if division:
        cpan_nodes = cpan_nodes.filter(te__nwo=division)
        maan_nodes = maan_nodes.filter(te__nwo=division)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    division_label = division.name if division else "ALL"
    ws.append(["Division", division_label])
    ws.append([])
    ws.append(["Metric", "Value"])

    metrics = [
        ("Total BTS", bts_list.count()),
        ("Cable Types", cable_summary.count()),
        ("Total Cables", cable_queryset.count()),
        ("Total Equipment", equipment_queryset.count()),
        ("Total Circuits", circuit_queryset.count()),
        ("CPAN Nodes", cpan_nodes.count()),
        ("MAAN Nodes", maan_nodes.count()),
    ]
    for k, v in metrics:
        ws.append([k, v])

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    ws_cable_types = wb.create_sheet("Cable Types")
    ws_cable_types.append(["Cable Type", "Mode", "Total"])
    for row in cable_summary:
        ws_cable_types.append([row.get("cable_type"), row.get("mode"), row.get("total")])
    ws_cable_types.freeze_panes = "A2"

    ws_oh = wb.create_sheet("Cables OH")
    ws_oh.append(["Cable Name", "TE", "Type", "Total Fibers", "Used", "Free"])
    for c in oh_cables.order_by("te__name", "name"):
        ws_oh.append([c.name, (c.te.name or "").upper(), c.cable_type, c.total_fibers, c.used_fibers, c.free_fibers])
    ws_oh.freeze_panes = "A2"

    ws_ug = wb.create_sheet("Cables UG")
    ws_ug.append(["Cable Name", "TE", "Type", "Total Fibers", "Used", "Free"])
    for c in ug_cables.order_by("te__name", "name"):
        ws_ug.append([c.name, (c.te.name or "").upper(), c.cable_type, c.total_fibers, c.used_fibers, c.free_fibers])
    ws_ug.freeze_panes = "A2"

    ws_equipment_types = wb.create_sheet("Equipment Types")
    ws_equipment_types.append(["Equipment Type", "Total"])
    for row in equipment_summary:
        ws_equipment_types.append([row.get("equipment_type"), row.get("total")])
    ws_equipment_types.freeze_panes = "A2"

    ws_cpan = wb.create_sheet("CPAN Nodes")
    ws_cpan.append(["Node Name", "TE", "Total Ports"])
    for eq in cpan_nodes.select_related("te").order_by("te__name", "name"):
        ws_cpan.append([eq.name, (eq.te.name or "").upper(), eq.total_ports])
    ws_cpan.freeze_panes = "A2"

    ws_maan = wb.create_sheet("MAAN Nodes")
    ws_maan.append(["Node Name", "Equipment Type", "TE", "Total Ports"])
    for eq in maan_nodes.select_related("te").order_by("te__name", "name"):
        ws_maan.append([eq.name, eq.equipment_type, (eq.te.name or "").upper(), eq.total_ports])
    ws_maan.freeze_panes = "A2"

    ws_circuits = wb.create_sheet("Circuit Types")
    ws_circuits.append(["Circuit Type", "Total"])
    for row in circuit_type_summary:
        ws_circuits.append([row.get("circuit_type"), row.get("total")])
    ws_circuits.freeze_panes = "A2"

    ws_bts = wb.create_sheet("BTS")
    ws_bts.append(["RP ID", "BTS Name", "TE", "Place", "Has CEF 12T", "Ring", "Latitude", "Longitude"])
    for b in bts_list.order_by("rp_id"):
        te_name = ""
        if b.maan_node and b.maan_node.te:
            te_name = (b.maan_node.te.name or "").upper()
        ws_bts.append([b.rp_id, b.bts_name, te_name, b.place_name or "", "Y" if b.has_cef_12t else "N", "Y" if b.is_ring else "N", b.latitude, b.longitude])
    ws_bts.freeze_panes = "A2"

    for sheet in wb.worksheets:
        max_col = sheet.max_column
        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = max(sheet.column_dimensions[column_letter].width or 10, 14)

    filename = f"analytics_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_cables(request):
    cables = Cable.objects.all()
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            cables = cables.filter(te__nwo=request.user.profile.division)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cables Report"
    
    columns = ['Name', 'Cable Type', 'Mode', 'TE Name', 'Remarks']
    ws.append(columns)
    
    for cable in cables:
        ws.append([
            cable.name,
            cable.cable_type,
            cable.mode,
            cable.te.name if cable.te else "N/A",
            cable.remarks or ""
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cables_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_equipment(request):
    equipment = Equipment.objects.all()
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            equipment = equipment.filter(te__nwo=request.user.profile.division)
            
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Report"
    
    columns = ['Name', 'Equipment Type', 'Total Ports', 'TE Name', 'Remarks']
    ws.append(columns)
    
    for eq in equipment:
        ws.append([
            eq.name,
            eq.equipment_type,
            eq.total_ports,
            eq.te.name if eq.te else "N/A",
            eq.remarks or ""
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="equipment_report.xlsx"'
    wb.save(response)
    return response

@login_required
def export_bts(request):
    queryset = MobileBTS.objects.filter(site_type='4G').select_related('maan_node', 'maan_node__te').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(Q(maan_node__te__nwo=division) | Q(te__nwo=division)).distinct()

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(Q(rp_id__icontains=search) | Q(bts_name__icontains=search))

    is_ring = request.GET.get('is_ring')
    if is_ring:
        queryset = queryset.filter(is_ring=is_ring == 'true')

    wb = Workbook()
    ws = wb.active
    ws.title = "BTS Report"

    columns = [
        'RP ID', 'BTS Name', 'MAAN Node', 'TE', 'Place Name',
        'Latitude', 'Longitude', 'Has CEF 12T', 'Is Ring', 'ERPS Image URL',
        'P2 Circuit', 'P2 System End', 'P2 Cable',
        'P3 Circuit', 'P3 System End', 'P3 Cable',
        'P4 Circuit', 'P4 System End', 'P4 Cable',
        'P5 Circuit', 'P5 System End', 'P5 Cable',
    ]
    ws.append(columns)

    for bts in queryset.order_by('rp_id'):
        ports = bts.cef_ports_data or {}
        def port_value(port_key, field_key):
            data = ports.get(port_key, {})
            return data.get(field_key, '') if isinstance(data, dict) else ''

        te_name = ''
        maan_node_name = ''
        if bts.maan_node:
            maan_node_name = bts.maan_node.name or ''
            if bts.maan_node.te:
                te_name = bts.maan_node.te.name or ''

        erps_url = ''
        if getattr(bts, 'erps_image', None):
            try:
                erps_url = bts.erps_image.url
            except Exception:
                erps_url = ''

        ws.append([
            bts.rp_id,
            bts.bts_name,
            maan_node_name,
            te_name,
            bts.place_name or '',
            float(bts.latitude) if bts.latitude is not None else '',
            float(bts.longitude) if bts.longitude is not None else '',
            'YES' if bts.has_cef_12t else 'NO',
            'YES' if bts.is_ring else 'NO',
            erps_url,
            port_value('p2', 'circuit'),
            port_value('p2', 'system_end'),
            port_value('p2', 'cable'),
            port_value('p3', 'circuit'),
            port_value('p3', 'system_end'),
            port_value('p3', 'cable'),
            port_value('p4', 'circuit'),
            port_value('p4', 'system_end'),
            port_value('p4', 'cable'),
            port_value('p5', 'circuit'),
            port_value('p5', 'system_end'),
            port_value('p5', 'cable'),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    division_label = division.name if division else 'all'
    filename = f"bts_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_no_4g_bts(request):
    queryset = MobileBTS.objects.filter(site_type='NON_4G').select_related('te', 'te__nwo').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(Q(maan_node__te__nwo=division) | Q(te__nwo=division)).distinct()

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(Q(rp_id__icontains=search) | Q(bts_name__icontains=search))

    wb = Workbook()
    ws = wb.active
    ws.title = "No 4G BTS Report"

    columns = [
        'RP ID', 'Site Name', 'Telephone Exchange', 'Place Name',
        'Latitude', 'Longitude', 'No 4G Site Type', 'Backhaul Media',
        'Connected Node/Equipment', 'Remarks'
    ]
    ws.append(columns)

    for bts in queryset.order_by('rp_id'):
        te_name = bts.te.name if bts.te else ''
        ws.append([
            bts.rp_id,
            bts.bts_name,
            te_name,
            bts.place_name or '',
            float(bts.latitude) if bts.latitude is not None else '',
            float(bts.longitude) if bts.longitude is not None else '',
            bts.get_non_4g_type_display() if bts.non_4g_type else '',
            bts.get_backhaul_media_display() if bts.backhaul_media else '',
            bts.connected_equipment or '',
            bts.remarks or ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    division_label = division.name if division else 'all'
    filename = f"no_4g_bts_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_ftth(request):
    queryset = FTTH.objects.select_related('division', 'te').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(division=division)

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(customer_name__icontains=search) |
            Q(landline_number__icontains=search) |
            Q(olt_name__icontains=search)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "FTTH Report"

    columns = [
        'Customer Name', 'Landline Number', 'Optical Power (dB)', 'OLT Name', 'Port Number',
        'Division', 'TE', 'Latitude', 'Longitude', 'Created At'
    ]
    ws.append(columns)

    for ftth in queryset.order_by('-created_at'):
        division_name = ''
        if ftth.division:
            try:
                division_name = ftth.division.get_name_display()
            except Exception:
                division_name = ftth.division.name

        te_name = ftth.te.name if ftth.te else ''

        ws.append([
            ftth.customer_name,
            ftth.landline_number,
            float(ftth.optical_power) if ftth.optical_power is not None else '',
            ftth.olt_name,
            ftth.port_number,
            division_name,
            te_name,
            float(ftth.latitude) if ftth.latitude is not None else '',
            float(ftth.longitude) if ftth.longitude is not None else '',
            ftth.created_at,
        ])

    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 10, 16)

    division_label = division.name if division else 'all'
    filename = f"ftth_report_{division_label.lower().replace(' ', '_')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_circuits(request):
    queryset = EBCircuit.objects.select_related('te', 'te__nwo', 'cable', 'equipment').all()
    division = None
    if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
        division = request.user.profile.division
        queryset = queryset.filter(te__nwo=division)

    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(client_name__icontains=search)

    circuit_type = request.GET.get('circuit_type')
    if circuit_type:
        queryset = queryset.filter(circuit_type=circuit_type)

    te_id = request.GET.get('te')
    if te_id:
        queryset = queryset.filter(te_id=te_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "EB Circuits"

    columns = [
        'SL No', 'TE Name', 'TYPE', 'NAME', 'LC ID', 'A-Media', 'Bandwidth', 'A-Address',
        'Node at A-End', 'Node at B-End', 'Port – B Side', 'Status', 'Working Status',
        'Fiber Mode', 'Cable Data', 'Customer Premise Location', 'OTDR Distance',
        'Latitude', 'Longitude', 'Remarks'
    ]
    ws.append(columns)

    for idx, c in enumerate(queryset.order_by('te__name', 'client_name'), start=1):
        te_name = c.te.name if c.te else ''
        ws.append([
            idx,
            te_name,
            c.circuit_type,
            c.client_name,
            c.lc_id or '',
            c.a_media or '',
            c.bandwidth,
            c.a_address or '',
            c.node_at_a_end or '',
            c.node_at_b_end or '',
            c.port_b_side or '',
            c.status or '',
            c.working_status or '',
            c.fiber_mode,
            c.cable_data or '',
            c.customer_premise_location or '',
            c.otdr_distance or '',
            c.latitude or '',
            c.longitude or '',
            c.remarks or '',
        ])

    ws.freeze_panes = "A2"
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 10, 16)

    division_label = division.name if division else 'all'
    filename = f"eb_circuits_{division_label.lower().replace(' ', '_')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response

def _resolve_te_helper(te_name, division=None):
    if not te_name:
        return None
    name_str = str(te_name).strip()
    
    import re
    # Substring spelling corrections to handle spelling variations within larger strings
    name_upper = name_str.upper()
    corrections = {
        'AYAPPANKAV': 'AYYAPPANKAVU',
        'CHITOOR': 'CHITTOOR',
        'CHITUR': 'CHITTOOR',
        'PANAMBILLI': 'PANAMPILLY',
        'PANAMPILLYNGR': 'PANAMPILLY NAGAR',
        'BOAT JETTY': 'BOATJETTY',
        'CARRIER STATION ROAD': 'CSR',
        'KADVNTRA': 'PANAMPILLYNAGAR',
        'KDVNTRA': 'PANAMPILLYNAGAR',
        'KADVNTHRA': 'PANAMPILLYNAGAR',
        'KADAVANTHRA': 'PANAMPILLYNAGAR',
        'PNR': 'PANAMPILLYNAGAR',
    }
    for search, replace in corrections.items():
        if search in name_upper:
            name_str = re.sub(re.escape(search), replace, name_str, flags=re.IGNORECASE)
            name_upper = name_str.upper()
            
    # Custom mapping dictionary for spelling variations and common abbreviations
    mapping = {
        'BOAT JETTY ERNAKULAM': 'Boatjetty TE',
        'BOAT JETTY': 'Boatjetty TE',
        'BOATJETTY': 'Boatjetty TE',
        'CARRIER STATION ROAD': 'Csr TE',
        'CSR': 'Csr TE',
        'KALOOR-ERNAKULAM': 'Kaloor TE',
        'KALOOR': 'Kaloor TE',
        'PANAMPILLYNAGAR': 'Panampilly Nagar TE',
        'PANAMPILLY NAGAR': 'Panampilly Nagar TE',
        'PANAMBILLI NAGAR': 'Panampilly Nagar TE',
        'PANAMPILLYNGR': 'Panampilly Nagar TE',
        'AYAPPANKAV': 'Ayyappankavu TE',
        'AYYAPPANKAVU': 'Ayyappankavu TE',
        'CHITOOR': 'Chittoor TE',
        'CHITTOOR': 'Chittoor TE',
        'CHITUR': 'Chittoor TE',
        'SRM ROAD': 'Srm TE',
        'SRM': 'Srm TE',
        'KADVNTRA': 'Panampilly Nagar TE',
        'KDVNTRA': 'Panampilly Nagar TE',
        'KADVNTHRA': 'Panampilly Nagar TE',
        'KADAVANTHRA': 'Panampilly Nagar TE',
        'PNR': 'Panampilly Nagar TE',
    }
    
    lookup_upper = name_str.upper().replace(' ', '').replace('-', '')
    mapping_normalized = {k.upper().replace(' ', '').replace('-', ''): v for k, v in mapping.items()}
    
    if lookup_upper in mapping_normalized:
        name_str = mapping_normalized[lookup_upper]
    
    # 1. Exact match
    qs = TelephoneExchange.objects.filter(name=name_str)
    if division:
        qs = qs.filter(nwo=division)
    te = qs.first()
    if te:
        return te
        
    # 2. Case-insensitive exact match
    qs = TelephoneExchange.objects.filter(name__iexact=name_str)
    if division:
        qs = qs.filter(nwo=division)
    te = qs.first()
    if te:
        return te
        
    # 3. Appending/removing suffix " TE"
    alt_name = name_str
    if alt_name.upper().endswith(" TE"):
        alt_name = alt_name[:-3].strip()
    else:
        alt_name = f"{alt_name} TE"
        
    qs = TelephoneExchange.objects.filter(name__iexact=alt_name)
    if division:
        qs = qs.filter(nwo=division)
    te = qs.first()
    if te:
        return te
        
    # 4. Fallback: case-insensitive contains match (only if it matches exactly 1 exchange)
    qs = TelephoneExchange.objects.filter(name__icontains=name_str)
    if division:
        qs = qs.filter(nwo=division)
    if qs.count() == 1:
        return qs.first()
        
    # 5. Reverse substring search: check if any exchange name (without " TE") is a substring of the lookup string
    lookup_clean = name_str.upper().replace(' ', '').replace('-', '')
    if len(lookup_clean) >= 3:
        all_exchanges = TelephoneExchange.objects.all()
        if division:
            all_exchanges = all_exchanges.filter(nwo=division)
        for exchange in all_exchanges:
            exch_clean = exchange.name.upper()
            if exch_clean.endswith(" TE"):
                exch_clean = exch_clean[:-3].strip()
            exch_clean = exch_clean.replace(' ', '').replace('-', '')
            if len(exch_clean) >= 3 and exch_clean in lookup_clean:
                return exchange
                
    return None

def _auto_assign_bts_helper(division):
    if not division:
        return
    unlinked_bts = MobileBTS.objects.filter(
        Q(te=None, maan_node=None) | 
        Q(te__name__startswith="UNMAPPED -", maan_node=None)
    )
    if unlinked_bts.exists():
        placeholder_name = f"UNMAPPED - {division.name}"
        placeholder_te, _ = TelephoneExchange.objects.get_or_create(
            name=placeholder_name,
            defaults={'nwo': division}
        )
        for bts in unlinked_bts:
            matched_te = None
            # 1. Try to match by place_name
            if bts.place_name:
                matched_te = _resolve_te_helper(bts.place_name, division)
            # 2. Try to match by bts_name
            if not matched_te and bts.bts_name:
                matched_te = _resolve_te_helper(bts.bts_name, division)
            
            # 3. Save matching or fallback to placeholder (only if te is currently None)
            if matched_te:
                if bts.te != matched_te:
                    bts.te = matched_te
                    bts.save()
            elif bts.te is None:
                bts.te = placeholder_te
                bts.save()

@login_required
def bulk_upload(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        upload_type = request.POST.get('upload_type')
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            header_map = {name: i for i, name in enumerate(headers)}
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            def get_value(row, *keys, required=False, default=None):
                for key in keys:
                    idx = header_map.get(key)
                    if idx is not None:
                        return row[idx]
                if required:
                    raise KeyError(f"Missing required column: {keys[0]}")
                return default

            def parse_bool(val):
                if val is None:
                    return False
                if isinstance(val, bool):
                    return val
                text = str(val).strip().lower()
                return text in {"y", "yes", "true", "1", "t"}

            created_count = 0
            skipped_count = 0
            row_errors = []
            
            if upload_type == 'CIRCUIT':
                for i, row in enumerate(rows, start=2):
                    try:
                        te_name = get_value(row, 'TE Name', 'TE')
                        division = None
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            division = request.user.profile.division
                        
                        te = None
                        if te_name not in (None, ''):
                            te = _resolve_te_helper(te_name, division)
                            
                        if not te:
                            if not division:
                                division = NWO.objects.first()
                            
                            placeholder_name = f"UNMAPPED - {division.name}" if division else "UNMAPPED - ALL"
                            te, _ = TelephoneExchange.objects.get_or_create(
                                name=placeholder_name,
                                defaults={'nwo': division} if division else {}
                            )
                            if te_name not in (None, ''):
                                row_errors.append(f"Row {i} (SL No {get_value(row, 'SL No', 'Sl No') or i-1}): Telephone Exchange '{te_name}' not found. Saved under '{placeholder_name}'.")
                            else:
                                row_errors.append(f"Row {i} (SL No {get_value(row, 'SL No', 'Sl No') or i-1}): Blank Telephone Exchange. Saved under '{placeholder_name}'.")

                        circuit_type_val = get_value(row, 'TYPE', 'Type')
                        if circuit_type_val not in (None, ''):
                            circuit_type_val = str(circuit_type_val).strip().upper()
                            type_mapping = {
                                'MPLS VPN LC': 'MPLS VPN',
                                'IINTERNET LC': 'INTERNET LC',
                                'INTERNET': 'INTERNET LC',
                                'P2P': 'P2P LC',
                            }
                            if circuit_type_val in type_mapping:
                                circuit_type_val = type_mapping[circuit_type_val]
                            valid_types = {'INTERNET LC', 'P2P LC', 'P2P LC ACROSS STATE', 'MPLS VPN', 'ISDN PRI'}
                            if circuit_type_val not in valid_types:
                                row_errors.append(f"Row {i} (SL No {get_value(row, 'SL No', 'Sl No') or i-1}): Invalid Circuit Type '{circuit_type_val}'. Set to blank.")
                                circuit_type_val = None
                        else:
                            circuit_type_val = None

                        # Map Node at A-End to DB Choices
                        end_node_raw = get_value(row, 'Node at A-End', 'End Node')
                        customer_end_node = None
                        if end_node_raw not in (None, ''):
                            end_node_raw = str(end_node_raw).strip().upper()
                            node_map = {
                                'CPE': 'CPE',
                                'MEDIA CONVERTER': 'MEDIA_CONVERTER',
                                'MEDIA-CONVERTER': 'MEDIA_CONVERTER',
                                'A-NODE': 'A_NODE',
                                'A_NODE': 'A_NODE',
                                'A NODE': 'A_NODE',
                                'CPAN B-NODE': 'CPAN_B',
                                'CPAN B': 'CPAN_B',
                                'CPAN_B': 'CPAN_B',
                                'MRO TEK': 'MRO_TEK',
                                'MRO-TEK': 'MRO_TEK',
                                'FTTH MODEM': 'FTTH_MODEM',
                                'FTTH-MODEM': 'FTTH_MODEM',
                                'MADM': 'MADM',
                                'MAAN A3 / A4': 'MAAN_A3_A4',
                                'MAAN A3/A4': 'MAAN_A3_A4',
                            }
                            customer_end_node = node_map.get(end_node_raw)

                        # Normalize Fiber Mode
                        fiber_mode_raw = get_value(row, 'Fiber Mode')
                        fiber_mode = None
                        if fiber_mode_raw not in (None, ''):
                            fiber_mode_raw = str(fiber_mode_raw).strip().upper()
                            if fiber_mode_raw in ('SINGLE', 'S', '1'):
                                fiber_mode = 'SINGLE'
                            elif fiber_mode_raw in ('DUAL', 'D', '2'):
                                fiber_mode = 'DUAL'
                            else:
                                row_errors.append(f"Row {i} (SL No {get_value(row, 'SL No', 'Sl No') or i-1}): Invalid Fiber Mode '{fiber_mode_raw}'. Set to blank.")

                        lat = get_value(row, 'Latitude')
                        lon = get_value(row, 'Longitude')
                        latitude = float(lat) if lat not in (None, '') else None
                        longitude = float(lon) if lon not in (None, '') else None

                        EBCircuit.objects.create(
                            te=te,
                            circuit_type=circuit_type_val,
                            client_name=get_value(row, 'NAME', 'Name', 'Client Name'),
                            lc_id=get_value(row, 'LC ID', 'lc_id'),
                            a_media=get_value(row, 'A-Media', 'A-media'),
                            bandwidth=get_value(row, 'Bandwidth'),
                            a_address=get_value(row, 'A-Address', 'A-address'),
                            node_at_a_end=get_value(row, 'Node at A-End'),
                            node_at_b_end=get_value(row, 'Node at B-End'),
                            port_b_side=get_value(row, 'Port – B Side', 'Port - B Side'),
                            status=get_value(row, 'Status'),
                            working_status=get_value(row, 'Working Status'),
                            customer_end_node=customer_end_node,
                            fiber_mode=fiber_mode,
                            cable_data=get_value(row, 'Cable Data'),
                            customer_premise_location=get_value(row, 'Customer Premise Location'),
                            otdr_distance=get_value(row, 'OTDR Distance'),
                            latitude=latitude,
                            longitude=longitude,
                            remarks=get_value(row, 'Remarks', default='Bulk Uploaded') or 'Bulk Uploaded'
                        )
                        created_count += 1
                    except IntegrityError as e:
                        row_errors.append(f"Row {i} (SL No {get_value(row, 'SL No', 'Sl No') or i-1}): Database integrity error - {e}")
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i} (SL No {get_value(row, 'SL No', 'Sl No') or i-1}): {e}")
                        skipped_count += 1
                if row_errors:
                    messages.warning(request, f"EB Circuits uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"EB Circuits uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")
            
            elif upload_type == 'CABLE':
                for i, row in enumerate(rows, start=2):
                    try:
                        te_name = get_value(row, 'TE', 'TE Name', required=True)
                        division = None
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            division = request.user.profile.division
                        te = _resolve_te_helper(te_name, division)
                        if not te:
                            row_errors.append(f"Row {i}: Telephone Exchange '{te_name}' not found.")
                            skipped_count += 1
                            continue

                        Cable.objects.create(
                            name=get_value(row, 'Cable Name', required=True),
                            cable_type=get_value(row, 'Type', required=True),
                            fiber_count=get_value(row, 'Fiber Count', required=True),
                            mode=get_value(row, 'Mode', required=True),
                            te=te,
                            remarks=get_value(row, 'Remarks', default='Bulk Uploaded') or 'Bulk Uploaded'
                        )
                        created_count += 1
                    except IntegrityError as e:
                        row_errors.append(f"Row {i}: Database integrity error - {e}")
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"Cables uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"Cables uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'EQUIPMENT':
                for i, row in enumerate(rows, start=2):
                    try:
                        te_name = get_value(row, 'TE', 'TE Name', required=True)
                        division = None
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            division = request.user.profile.division
                        te = _resolve_te_helper(te_name, division)
                        if not te:
                            row_errors.append(f"Row {i}: Telephone Exchange '{te_name}' not found.")
                            skipped_count += 1
                            continue

                        Equipment.objects.create(
                            name=get_value(row, 'Equipment Name', required=True),
                            equipment_type=get_value(row, 'Type', required=True),
                            total_ports=get_value(row, 'Total Ports', required=True),
                            te=te,
                            remarks=get_value(row, 'Remarks', default='Bulk Uploaded') or 'Bulk Uploaded'
                        )
                        created_count += 1
                    except IntegrityError as e:
                        row_errors.append(f"Row {i}: Database integrity error - {e}")
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"Equipment uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"Equipment uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'BTS':
                for i, row in enumerate(rows, start=2):
                    try:
                        rp_id = str(get_value(row, 'RP ID', required=True)).strip()
                        bts_name = get_value(row, 'BTS Name', required=True)
                        
                        te_name = get_value(row, 'TE Name', 'TE')
                        division = None
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            division = request.user.profile.division
                        
                        te = None
                        if te_name not in (None, ''):
                            te = _resolve_te_helper(te_name, division)
                            
                        if not te:
                            if not division:
                                division = NWO.objects.first()
                            
                            placeholder_name = f"UNMAPPED - {division.name}" if division else "UNMAPPED - ALL"
                            te, _ = TelephoneExchange.objects.get_or_create(
                                name=placeholder_name,
                                defaults={'nwo': division} if division else {}
                            )
                            if te_name not in (None, ''):
                                row_errors.append(f"Row {i}: Telephone Exchange '{te_name}' not found. Saved under '{placeholder_name}'.")

                        place_name = get_value(row, 'Place Name')
                        latitude = get_value(row, 'Latitude')
                        longitude = get_value(row, 'Longitude')
                        has_cef_12t = parse_bool(get_value(row, 'Has CEF 12T', default=False))
                        is_ring = parse_bool(get_value(row, 'Is Ring', default=False))

                        lat_dec = None
                        if latitude not in (None, ''):
                            try:
                                lat_dec = Decimal(str(latitude))
                            except (InvalidOperation, ValueError, TypeError) as e:
                                row_errors.append(f"Row {i}: Invalid Latitude '{latitude}'. Set to blank.")
                        
                        lon_dec = None
                        if longitude not in (None, ''):
                            try:
                                lon_dec = Decimal(str(longitude))
                            except (InvalidOperation, ValueError, TypeError) as e:
                                row_errors.append(f"Row {i}: Invalid Longitude '{longitude}'. Set to blank.")

                        ports_data = {}
                        for port in ['P2', 'P3', 'P4', 'P5']:
                            ports_data[port.lower()] = {
                                'circuit': get_value(row, f'{port} Circuit', default='') or '',
                                'system_end': get_value(row, f'{port} System End', default='') or '',
                                'cable': get_value(row, f'{port} Cable', default='') or '',
                            }

                        MobileBTS.objects.update_or_create(
                            rp_id=rp_id,
                            defaults={
                                'site_type': '4G',
                                'te': te,
                                'bts_name': bts_name,
                                'place_name': place_name,
                                'latitude': lat_dec,
                                'longitude': lon_dec,
                                'has_cef_12t': has_cef_12t,
                                'is_ring': is_ring,
                                'cef_ports_data': ports_data,
                            }
                        )
                        created_count += 1
                    except IntegrityError as e:
                        row_errors.append(f"Row {i}: Database integrity error - {e}")
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                        skipped_count += 1
                
                from django.utils.safestring import mark_safe
                if row_errors:
                    msg = f"BTS uploaded with {len(row_errors)} warnings/errors. Created/Updated: {created_count}, Skipped: {skipped_count}."
                    msg += "<br><ul class='mb-0 ps-3'>" + "".join([f"<li>{err}</li>" for err in row_errors[:15]]) + ("<li>...and more</li>" if len(row_errors) > 15 else "") + "</ul>"
                    messages.warning(request, mark_safe(msg))
                else:
                    messages.success(request, f"BTS uploaded successfully! Created/Updated: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'BTS_NO_4G':
                for i, row in enumerate(rows, start=2):
                    try:
                        rp_id = str(get_value(row, 'RP ID', required=True)).strip()
                        bts_name = get_value(row, 'Site Name', 'BTS Name', 'Name', required=True)
                        te_name = get_value(row, 'TE', 'TE Name', required=True)
                        
                        division = None
                        if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                            division = request.user.profile.division
                        te = _resolve_te_helper(te_name, division)
                        if not te:
                            if not division:
                                division = NWO.objects.first()
                            
                            placeholder_name = f"UNMAPPED - {division.name}" if division else "UNMAPPED - ALL"
                            te, _ = TelephoneExchange.objects.get_or_create(
                                name=placeholder_name,
                                defaults={'nwo': division} if division else {}
                            )
                            if te_name not in (None, ''):
                                row_errors.append(f"Row {i}: Telephone Exchange '{te_name}' not found. Saved under '{placeholder_name}'.")

                        place_name = get_value(row, 'Place Name')
                        latitude = get_value(row, 'Latitude')
                        longitude = get_value(row, 'Longitude')
                        non_4g_type_val = get_value(row, 'No 4G Site Type', 'Site Type')
                        backhaul_media_val = get_value(row, 'Backhaul Media')
                        connected_equipment = get_value(row, 'Connected Node/Equipment', 'Connected Node')
                        remarks = get_value(row, 'Remarks')

                        # Normalize backhaul media choice
                        if backhaul_media_val not in (None, ''):
                            bm_upper = str(backhaul_media_val).strip().upper()
                            if bm_upper in ('FIBER', 'FIBRE', 'F'):
                                backhaul_media_val = 'FIBER'
                            elif bm_upper in ('MICROWAVE', 'MW', 'M'):
                                backhaul_media_val = 'MICROWAVE'
                            elif bm_upper in ('LEASED LINE', 'LEASED_LINE', 'LL', 'L'):
                                backhaul_media_val = 'LEASED_LINE'
                            else:
                                row_errors.append(f"Row {i}: Invalid Backhaul Media '{backhaul_media_val}'. Set to blank.")
                                backhaul_media_val = None
                        else:
                            backhaul_media_val = None

                        # Normalize non-4G type choice
                        if non_4g_type_val not in (None, ''):
                            nt_upper = str(non_4g_type_val).strip().upper().replace(' ', '').replace('+', '_')
                            type_map = {
                                '2G': '2G', '3G': '3G', '5G': '5G',
                                '2G_3G': '2G_3G', '3G_5G': '3G_5G', '2G_3G_5G': '2G_3G_5G',
                                '2G3G': '2G_3G', '3G5G': '3G_5G', '2G3G5G': '2G_3G_5G'
                            }
                            non_4g_type_val = type_map.get(nt_upper)
                            if not non_4g_type_val:
                                row_errors.append(f"Row {i}: Invalid No 4G Site Type '{get_value(row, 'No 4G Site Type', 'Site Type')}'. Set to blank.")
                        else:
                            non_4g_type_val = None

                        lat_dec = None
                        if latitude not in (None, ''):
                            try:
                                lat_dec = Decimal(str(latitude))
                            except (InvalidOperation, ValueError, TypeError) as e:
                                row_errors.append(f"Row {i}: Invalid Latitude '{latitude}'. Set to blank.")
                        
                        lon_dec = None
                        if longitude not in (None, ''):
                            try:
                                lon_dec = Decimal(str(longitude))
                            except (InvalidOperation, ValueError, TypeError) as e:
                                row_errors.append(f"Row {i}: Invalid Longitude '{longitude}'. Set to blank.")

                        MobileBTS.objects.update_or_create(
                            rp_id=rp_id,
                            defaults={
                                'site_type': 'NON_4G',
                                'bts_name': bts_name,
                                'te': te,
                                'place_name': place_name,
                                'latitude': lat_dec,
                                'longitude': lon_dec,
                                'non_4g_type': non_4g_type_val,
                                'backhaul_media': backhaul_media_val,
                                'connected_equipment': connected_equipment,
                                'remarks': remarks
                            }
                        )
                        created_count += 1
                    except IntegrityError as e:
                        row_errors.append(f"Row {i}: Database integrity error - {e}")
                        skipped_count += 1
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                        skipped_count += 1
                
                from django.utils.safestring import mark_safe
                if row_errors:
                    msg = f"No 4G BTS uploaded with {len(row_errors)} warnings/errors. Created/Updated: {created_count}, Skipped: {skipped_count}."
                    msg += "<br><ul class='mb-0 ps-3'>" + "".join([f"<li>{err}</li>" for err in row_errors[:15]]) + ("<li>...and more</li>" if len(row_errors) > 15 else "") + "</ul>"
                    messages.warning(request, mark_safe(msg))
                else:
                    messages.success(request, f"No 4G BTS uploaded successfully! Created/Updated: {created_count}, Skipped: {skipped_count}.")

            elif upload_type == 'FTTH':
                user_division = None
                if not request.user.is_superuser and hasattr(request.user, 'profile') and request.user.profile.division:
                    user_division = request.user.profile.division

                for i, row in enumerate(rows, start=2):
                    try:
                        customer_name = get_value(row, 'Customer Name', required=True)
                        landline_number = str(get_value(row, 'Landline Number', required=True)).strip()
                        optical_power_raw = get_value(row, 'Optical Power', required=True)
                        olt_name = get_value(row, 'OLT Name', required=True)
                        port_number = int(get_value(row, 'Port Number', required=True))
                        division_name = get_value(row, 'Division')
                        te_name = get_value(row, 'TE', 'TE Name')
                        latitude = get_value(row, 'Latitude')
                        longitude = get_value(row, 'Longitude')

                        optical_power_str = str(optical_power_raw).strip()
                        optical_power_value = Decimal(optical_power_str.replace('dB', '').replace('DB', '').strip())

                        division = user_division
                        if division is None and division_name not in (None, ''):
                            division = NWO.objects.filter(name=str(division_name).strip()).first()

                        te = None
                        if te_name not in (None, ''):
                            te = _resolve_te_helper(te_name, division)
                            if not te:
                                row_errors.append(f"Row {i}: Telephone Exchange '{te_name}' not found.")
                                skipped_count += 1
                                continue
                            if division is None and te is not None:
                                division = te.nwo

                        FTTH.objects.create(
                            customer_name=customer_name,
                            landline_number=landline_number,
                            optical_power=optical_power_value,
                            olt_name=olt_name,
                            port_number=port_number,
                            division=division,
                            te=te,
                            latitude=Decimal(str(latitude)) if latitude not in (None, '') else None,
                            longitude=Decimal(str(longitude)) if longitude not in (None, '') else None,
                        )
                        created_count += 1
                    except IntegrityError as e:
                        row_errors.append(f"Row {i}: Database integrity error - {e}")
                        skipped_count += 1
                    except (InvalidOperation, ValueError) as e:
                        row_errors.append(f"Row {i}: Invalid numeric value ({e})")
                    except Exception as e:
                        row_errors.append(f"Row {i}: {e}")
                if row_errors:
                    messages.warning(request, f"FTTH uploaded with {len(row_errors)} row errors. Created: {created_count}, Skipped: {skipped_count}.")
                else:
                    messages.success(request, f"FTTH uploaded successfully! Created: {created_count}, Skipped: {skipped_count}.")
            
        except Exception as e:
            error_msg = str(e)
            if "not a zip file" in error_msg.lower():
                messages.error(request, "Error processing file: The uploaded file is not a valid Excel Workbook (.xlsx) file. If you are using an older Excel format (.xls) or CSV, please open the file and 'Save As' an 'Excel Workbook (*.xlsx)' before uploading.")
            else:
                messages.error(request, f"Error processing file: {error_msg}")
            
    return render(request, 'inventory/bulk_upload.html')

@login_required
def download_template(request):
    category = request.GET.get('category', 'CIRCUIT')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{category} Template"
    
    if category == 'CIRCUIT':
        columns = [
            'SL No', 'TE Name', 'TYPE', 'NAME', 'LC ID', 'A-Media', 'Bandwidth', 'A-Address',
            'Node at A-End', 'Node at B-End', 'Port – B Side', 'Status', 'Working Status',
            'Fiber Mode', 'Cable Data', 'Customer Premise Location', 'OTDR Distance',
            'Latitude', 'Longitude', 'Remarks'
        ]
        circuit_type = request.GET.get('circuit_type', 'INTERNET LC')
        valid_types = {'INTERNET LC', 'P2P LC', 'P2P LC ACROSS STATE', 'MPLS VPN', 'ISDN PRI'}
        if circuit_type not in valid_types:
            circuit_type = 'INTERNET LC'
            
        data = [
            1, 'Panambilly Nagar', circuit_type, 'Reliance JIO', 'LC-9988-KCH', 'Media Converter',
            '100 Mbps', 'Building A, Kochi', 'CPE', 'CPAN_B', 'GigabitEthernet1/1/1', 'Working',
            'Active', 'DUAL', 'PN-CSR-48F-01', 'Server Room, Ground Floor', '2.5 km',
            9.9816, 76.2999, 'Sample EB circuit'
        ]
    elif category == 'CABLE':
        columns = ['Cable Name', 'Type', 'Fiber Count', 'Mode', 'TE', 'Remarks']
        data = ['PN-CSR-48F-01', '48F', 48, 'UG', 'Panambilly Nagar', 'Sample cable']
    elif category == 'EQUIPMENT':
        columns = ['Equipment Name', 'Type', 'Total Ports', 'TE', 'Remarks']
        data = ['PN-CPAN-01', 'CPAN_B', 24, 'Panambilly Nagar', 'Sample equipment']
    elif category == 'BTS':
        columns = [
            'RP ID', 'BTS Name', 'TE Name', 'Place Name', 'Latitude', 'Longitude', 'Has CEF 12T', 'Is Ring',
            'P2 Circuit', 'P2 System End', 'P2 Cable',
            'P3 Circuit', 'P3 System End', 'P3 Cable',
            'P4 Circuit', 'P4 System End', 'P4 Cable',
            'P5 Circuit', 'P5 System End', 'P5 Cable',
        ]
        data = [
            'RP-0001', 'BTS Example Site', 'Panambilly Nagar', 'Sample Location', 9.931233, 76.267303, 'Y', 'N',
            'CIR-001', 'SYSTEM-A', 'CABLE-001',
            '', '', '',
            '', '', '',
            '', '', '',
        ]
    elif category == 'BTS_NO_4G':
        columns = [
            'RP ID', 'Site Name', 'TE Name', 'Place Name', 'Latitude', 'Longitude',
            'No 4G Site Type', 'Backhaul Media', 'Connected Node/Equipment', 'Remarks'
        ]
        data = [
            'RP-0001', 'BTS Example Site', 'Panambilly Nagar', 'Sample Location', 9.931233, 76.267303,
            '2G + 3G', 'Fiber', 'BSC', 'Sample No 4G Site'
        ]
    elif category == 'FTTH':
        columns = ['Customer Name', 'Landline Number', 'Optical Power', 'OLT Name', 'Port Number', 'Division', 'TE', 'Latitude', 'Longitude']
        data = ['Customer A', '0484XXXXXXX', '-23.45', 'OLT-1', 1, 'NWO KOCHI', 'Panambilly Nagar', 9.931233, 76.267303]
    else:
        columns = [
            'SL No', 'TE Name', 'TYPE', 'NAME', 'LC ID', 'A-Media', 'Bandwidth', 'A-Address',
            'Node at A-End', 'Node at B-End', 'Port – B Side', 'Status', 'Working Status',
            'Fiber Mode', 'Cable Data', 'Customer Premise Location', 'OTDR Distance',
            'Latitude', 'Longitude', 'Remarks'
        ]
        circuit_type = request.GET.get('circuit_type', 'INTERNET LC')
        valid_types = {'INTERNET LC', 'P2P LC', 'P2P LC ACROSS STATE', 'MPLS VPN', 'ISDN PRI'}
        if circuit_type not in valid_types:
            circuit_type = 'INTERNET LC'
            
        data = [
            1, 'Panambilly Nagar', circuit_type, 'Reliance JIO', 'LC-9988-KCH', 'Media Converter',
            '100 Mbps', 'Building A, Kochi', 'CPE', 'CPAN_B', 'GigabitEthernet1/1/1', 'Working',
            'Active', 'DUAL', 'PN-CSR-48F-01', 'Server Room, Ground Floor', '2.5 km',
            9.9816, 76.2999, 'Sample EB circuit'
        ]

        
    ws.append(columns)
    ws.append(data)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{category.lower()}_template.xlsx"'
    wb.save(response)
    return response

# OH Maintenance Module Views

class OHMaintenanceListView(LoginRequiredMixin, DivisionRequiredMixin, ListView):
    model = OHMaintenanceEntry
    template_name = 'inventory/oh_maintenance_list.html'
    context_object_name = 'entries'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        # Search filters
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(team_name__icontains(q)) |
                Q(work_order_no__icontains(q)) |
                Q(location__icontains(q)) |
                Q(route_name__icontains(q))
            )
        
        # Date range filter
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            queryset = queryset.filter(maintenance_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(maintenance_date__lte=end_date)
            
        # TE filter
        te_id = self.request.GET.get('te')
        if te_id:
            queryset = queryset.filter(te_id=te_id)
            
        return queryset

class OHMaintenanceCreateView(LoginRequiredMixin, DivisionRequiredMixin, CreateView):
    model = OHMaintenanceEntry
    form_class = OHMaintenanceEntryForm
    template_name = 'inventory/oh_maintenance_form.html'
    success_url = reverse_lazy('oh_maintenance_list')

    def get_initial(self):
        initial = super().get_initial()
        te_id = self.request.GET.get('te')
        if te_id:
            initial['te'] = te_id
        return initial

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['activities'] = OHMaintenanceActivityFormSet(self.request.POST)
        else:
            data['activities'] = OHMaintenanceActivityFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        activities = context['activities']
        
        # Auto-fetch division from profile
        if hasattr(self.request.user, 'profile') and self.request.user.profile.division:
            form.instance.division = self.request.user.profile.division
        
        form.instance.created_by = self.request.user
        
        if activities.is_valid():
            self.object = form.save()
            activities.instance = self.object
            activities.save()
            messages.success(self.request, "Maintenance entry created successfully.")
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class OHMaintenanceUpdateView(LoginRequiredMixin, DivisionRequiredMixin, UpdateView):
    model = OHMaintenanceEntry
    form_class = OHMaintenanceEntryForm
    template_name = 'inventory/oh_maintenance_form.html'
    success_url = reverse_lazy('oh_maintenance_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['activities'] = OHMaintenanceActivityFormSet(self.request.POST, instance=self.object)
        else:
            data['activities'] = OHMaintenanceActivityFormSet(instance=self.object)
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        activities = context['activities']
        
        form.instance.modified_by = self.request.user
        
        if activities.is_valid():
            self.object = form.save()
            activities.instance = self.object
            activities.save()
            messages.success(self.request, "Maintenance entry updated successfully.")
            return redirect(self.success_url)
        else:
            return self.render_to_response(self.get_context_data(form=form))

class OHMaintenanceDetailView(LoginRequiredMixin, DivisionRequiredMixin, DetailView):
    model = OHMaintenanceEntry
    template_name = 'inventory/oh_maintenance_detail.html'
    context_object_name = 'entry'

class OHMaintenanceDeleteView(LoginRequiredMixin, DivisionRequiredMixin, DeleteView):
    model = OHMaintenanceEntry
    template_name = 'inventory/oh_maintenance_confirm_delete.html'
    success_url = reverse_lazy('oh_maintenance_list')

@login_required
def get_activity_rates(request):
    rates = OHMaintenanceRateMaster.objects.filter(is_active=True).values(
        'activity_type', 'unit_type', 'unit_rate'
    )
    return JsonResponse(list(rates), safe=False)

class OHMaintenanceRateMasterListView(LoginRequiredMixin, ListView):
    model = OHMaintenanceRateMaster
    template_name = 'inventory/oh_rate_master_list.html'
    context_object_name = 'rates'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "Only admins can access Rate Master.")
            return redirect('oh_maintenance_list')
        return super().dispatch(request, *args, **kwargs)

class OHMaintenanceRateMasterCreateView(LoginRequiredMixin, CreateView):
    model = OHMaintenanceRateMaster
    fields = ['activity_type', 'unit_type', 'unit_rate', 'effective_from', 'effective_to', 'is_active']
    template_name = 'inventory/oh_rate_master_form.html'
    success_url = reverse_lazy('oh_rate_master_list')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "Only admins can access Rate Master.")
            return redirect('oh_maintenance_list')
        return super().dispatch(request, *args, **kwargs)

@login_required
def oh_maintenance_bill(request):
    division = None
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division
    else:
        division_id = request.GET.get('division')
        if division_id:
            division = NWO.objects.filter(id=division_id).first()

    month = request.GET.get('month', timezone.now().month)
    year = request.GET.get('year', timezone.now().year)
    
    entries = OHMaintenanceEntry.objects.filter(
        status='Approved',
        maintenance_date__month=month,
        maintenance_date__year=year
    )
    if division:
        entries = entries.filter(division=division)
        
    # Aggregate by activity type
    bill_data = entries.values('activities__activity_type', 'activities__unit_type', 'activities__unit_rate').annotate(
        total_quantity=Sum('activities__quantity'),
        total_amount=Sum('activities__amount')
    ).order_by('activities__activity_type')
    
    total_bill_amount = sum(item['total_amount'] or 0 for item in bill_data)

    context = {
        'bill_data': bill_data,
        'total_bill_amount': total_bill_amount,
        'month': month,
        'year': year,
        'division': division,
        'divisions': NWO.objects.all() if request.user.is_superuser else None
    }
    return render(request, 'inventory/oh_maintenance_bill.html', context)

@login_required
def oh_maintenance_dashboard(request):
    from django.db.models import F, ExpressionWrapper, DurationField, Avg, Count, Sum
    from django.utils import timezone
    from datetime import timedelta
    
    division = None
    if not request.user.is_superuser:
        if hasattr(request.user, 'profile') and request.user.profile.division:
            division = request.user.profile.division
    else:
        div_id = request.GET.get('division')
        if div_id:
            division = NWO.objects.filter(id=div_id).first()
            
    entries = OHMaintenanceEntry.objects.all()
    if division:
        entries = entries.filter(division=division)
    
    # Fault Tracking Stats
    now = timezone.now()
    first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculate durations for MTTR
    entries_with_duration = entries.filter(
        fault_occurrence_datetime__isnull=False,
        fault_clearance_datetime__isnull=False
    ).annotate(
        duration=ExpressionWrapper(
            F('fault_clearance_datetime') - F('fault_occurrence_datetime'),
            output_field=DurationField()
        )
    )
    
    avg_mttr = entries_with_duration.aggregate(avg=Avg('duration'))['avg']
    
    stats = {
        'total_entries': entries.count(),
        'pending_verification': entries.filter(status='Submitted').count(),
        'pending_approval': entries.filter(status='Verified').count(),
        'total_cost_this_month': entries.filter(
            maintenance_date__month=now.month,
            maintenance_date__year=now.year
        ).aggregate(total=Sum('activities__amount'))['total'] or 0,
        
        # New Fault Tracking Stats
        'total_faults_this_month': entries.filter(
            fault_occurrence_datetime__gte=first_day_of_month
        ).count(),
        'avg_clearance_time': avg_mttr,
        'critical_fault_count': entries.filter(fault_severity='Critical').count(),
        'pending_fault_closures': entries.filter(
            fault_occurrence_datetime__isnull=False,
            fault_clearance_datetime__isnull=True
        ).count(),
    }
    
    # Division-wise MTTR for Graph
    division_mttr = entries_with_duration.values('division__name').annotate(
        avg_duration=Avg('duration')
    ).order_by('division__name')
    
    # Convert timedelta to hours for display in graph
    for item in division_mttr:
        if item['avg_duration']:
            item['avg_hours'] = item['avg_duration'].total_seconds() / 3600
        else:
            item['avg_hours'] = 0

    # Repeated Fault Hotspot Locations
    hotspots = entries.values('location').annotate(
        fault_count=Count('id')
    ).filter(fault_count__gt=1).order_by('-fault_count')[:10]
    
    # Activity-wise summary for chart
    activity_summary = OHMaintenanceActivity.objects.filter(entry__in=entries).values('activity_type').annotate(
        total_qty=Sum('quantity')
    ).order_by('-total_qty')

    context = {
        'stats': stats,
        'activity_summary': activity_summary,
        'recent_entries': entries[:10],
        'division_mttr': division_mttr,
        'hotspots': hotspots,
        'divisions': NWO.objects.all() if request.user.is_superuser else None,
        'selected_division': division,
    }
    return render(request, 'inventory/oh_maintenance_dashboard.html', context)


# ==============================================================================
# DATABASE BACKUP & RESTORE FACILITY (SUPERUSERS ONLY)
# ==============================================================================
from django.conf import settings
from django.db import transaction
from django.core.management import call_command
from django.http import HttpResponseForbidden
from django.contrib.auth.models import User
import tarfile
import io
import os
import shutil
import tempfile

@login_required
def db_management(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Only superusers are allowed to access Database Management.")
    
    # Calculate media folder size
    media_size = 0
    media_count = 0
    if os.path.exists(settings.MEDIA_ROOT):
        for root, dirs, files in os.walk(settings.MEDIA_ROOT):
            for file in files:
                media_size += os.path.getsize(os.path.join(root, file))
                media_count += 1
                
    # Calculate database counts
    stats = {
        'divisions': NWO.objects.count(),
        'exchanges': TelephoneExchange.objects.count(),
        'cables': Cable.objects.count(),
        'equipment': Equipment.objects.count(),
        'circuits': EBCircuit.objects.count(),
        'bts': MobileBTS.objects.count(),
        'jbs': JunctionBox.objects.count(),
        'lius': LIU.objects.count(),
        'ftth': FTTH.objects.count(),
        'splicings': Splicing.objects.count(),
        'users': User.objects.count(),
        'media_size_mb': round(media_size / (1024 * 1024), 2),
        'media_count': media_count,
    }
    
    return render(request, 'inventory/db_management.html', {'stats': stats})

@login_required
def db_backup(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Only superusers are allowed to download backups.")
        
    try:
        # Create an in-memory tar archive
        tar_buffer = io.BytesIO()
        with tarfile.open(fileobj=tar_buffer, mode="w:gz") as tar:
            # 1. Run dumpdata for database records
            db_buffer = io.StringIO()
            call_command(
                'dumpdata', 
                indent=2, 
                stdout=db_buffer, 
                exclude=['contenttypes', 'auth.Permission', 'admin.LogEntry', 'sessions.Session']
            )
            db_json = db_buffer.getvalue().encode('utf-8')
            
            # Add database dump to tar
            json_info = tarfile.TarInfo(name="db_backup.json")
            json_info.size = len(db_json)
            tar.addfile(json_info, io.BytesIO(db_json))
            
            # 2. Add media files to tar
            if os.path.exists(settings.MEDIA_ROOT):
                for root, dirs, files in os.walk(settings.MEDIA_ROOT):
                    for file in files:
                        full_path = os.path.join(root, file)
                        rel_path = os.path.relpath(full_path, settings.MEDIA_ROOT)
                        tar_path = os.path.join("media", rel_path)
                        tar.add(full_path, arcname=tar_path)
                        
        response = HttpResponse(tar_buffer.getvalue(), content_type='application/x-gzip')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=\"ofcnet_backup_{timestamp}.tar.gz\"'
        return response
    except Exception as e:
        messages.error(request, f"Backup creation failed: {str(e)}")
        return redirect('db_management')

@login_required
@require_http_methods(["POST"])
def db_restore(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Only superusers are allowed to restore database backups.")
        
    uploaded_file = request.FILES.get('backup_file')
    if not uploaded_file:
        messages.error(request, "Please select a backup file to upload.")
        return redirect('db_management')
        
    if not uploaded_file.name.endswith('.tar.gz') and not uploaded_file.name.endswith('.tgz'):
        messages.error(request, "Invalid file format. Please upload a compressed (.tar.gz) backup archive.")
        return redirect('db_management')
        
    # Write uploaded data to a temporary file
    with tempfile.NamedTemporaryFile(suffix='.tar.gz', delete=False) as temp_file:
        for chunk in uploaded_file.chunks():
            temp_file.write(chunk)
        temp_file_path = temp_file.name
        
    try:
        with tarfile.open(temp_file_path, mode="r:gz") as tar:
            # Check for db_backup.json inside the archive
            try:
                db_member = tar.getmember("db_backup.json")
            except KeyError:
                messages.error(request, "Invalid backup archive: 'db_backup.json' is missing.")
                return redirect('db_management')
                
            # Perform database restoration inside atomic transaction
            with transaction.atomic():
                # Delete existing database records in reverse dependency order
                Splicing.objects.all().delete()
                FTTH.objects.all().delete()
                MobileBTS.objects.all().delete()
                EBCircuit.objects.all().delete()
                LIU.objects.all().delete()
                JunctionBox.objects.all().delete()
                Equipment.objects.all().delete()
                Cable.objects.all().delete()
                TelephoneExchange.objects.all().delete()
                
                # Exclude active admin user to avoid locking ourselves out during restore
                UserProfile.objects.exclude(user=request.user).delete()
                User.objects.exclude(id=request.user.id).delete()
                NWO.objects.all().delete()
                
                # Extract and load database records
                db_file = tar.extractfile(db_member)
                with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as db_temp:
                    db_temp.write(db_file.read())
                    db_temp_path = db_temp.name
                    
                try:
                    call_command('loaddata', db_temp_path)
                finally:
                    os.remove(db_temp_path)
            
            # Extract and restore media files
            # Clean up the current media directory first to prevent orphaned media files
            if os.path.exists(settings.MEDIA_ROOT):
                shutil.rmtree(settings.MEDIA_ROOT)
            os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
            
            for member in tar.getmembers():
                if member.name.startswith("media/"):
                    target_rel_path = os.path.relpath(member.name, "media")
                    target_full_path = os.path.join(settings.MEDIA_ROOT, target_rel_path)
                    
                    if member.isdir():
                        os.makedirs(target_full_path, exist_ok=True)
                    else:
                        os.makedirs(os.path.dirname(target_full_path), exist_ok=True)
                        with tar.extractfile(member) as source_file, open(target_full_path, "wb") as dest_file:
                            shutil.copyfileobj(source_file, dest_file)
                            
            messages.success(request, "Database and media archive successfully restored!")
    except Exception as e:
        messages.error(request, f"Restore failed: {str(e)}")
    finally:
        if os.path.exists(temp_file_path):
            os.remove(temp_file_path)
            
    return redirect('db_management')

def debug_db(request):
    import traceback
    from django.db import connection
    from django.http import HttpResponse
    from django.core.management import call_command
    import os
    res = []
    
    db_url = os.environ.get('DATABASE_URL', 'Not Set')
    masked_url = db_url
    if '@' in db_url:
        parts = db_url.split('@')
        masked_url = "postgres://***:***@" + parts[-1]
    res.append("Database URL: " + masked_url)
    
    try:
        connection.ensure_connection()
        res.append("Database connection: SUCCESS")
        
        tables = connection.introspection.table_names()
        res.append(f"Tables in database ({len(tables)}):")
        for t in tables:
            res.append(f" - {t}")
            
        # Test session creation
        from django.contrib.sessions.backends.db import SessionStore
        try:
            s = SessionStore()
            s['test_key'] = 'test_value'
            s.create()
            res.append(f"\nSession creation test: SUCCESS (session_key={s.session_key})")
        except Exception as sexc:
            res.append(f"\nSession creation test: FAILED: {str(sexc)}")
            res.append(traceback.format_exc())
            
        # Log settings.DEBUG status
        from django.conf import settings as django_settings
        res.append(f"\nDEBUG setting status: {django_settings.DEBUG}")
        
        # Test authentication
        from django.contrib.auth import authenticate
        try:
            user = authenticate(username='nwo_ekm', password='Nwo#Ekm@2026!')
            if user is not None:
                res.append(f"Auth test for nwo_ekm: SUCCESS (superuser={user.is_superuser})")
            else:
                res.append("Auth test for nwo_ekm: FAILED (returned None)")
        except Exception as auth_exc:
            res.append(f"Auth test for nwo_ekm: FAILED with exception: {str(auth_exc)}")
            res.append(traceback.format_exc())
            
        # Test login flow
        from django.contrib.auth import login
        from django.test import RequestFactory
        try:
            factory = RequestFactory()
            dummy_request = factory.post('/')
            from django.contrib.sessions.middleware import SessionMiddleware
            middleware = SessionMiddleware(lambda req: HttpResponse())
            middleware(dummy_request)
            dummy_request.session.save()
            
            user = authenticate(username='nwo_ekm', password='Nwo#Ekm@2026!')
            if user:
                login(dummy_request, user)
                res.append(f"Login test for nwo_ekm: SUCCESS, session user_id={dummy_request.session.get('_auth_user_id')}")
            else:
                res.append("Login test for nwo_ekm: skipped (auth returned None)")
        except Exception as login_exc:
            res.append(f"Login test for nwo_ekm: FAILED with exception: {str(login_exc)}")
            res.append(traceback.format_exc())
            
        from django.db.migrations.recorder import MigrationRecorder
        recorder = MigrationRecorder(connection)
        applied_migs = recorder.applied_migrations()
        res.append(f"\nApplied migrations count: {len(applied_migs)}")
        
        from django.db.migrations.executor import MigrationExecutor
        executor = MigrationExecutor(connection)
        plan = executor.migration_plan(executor.loader.graph.leaf_nodes())
        res.append(f"Unapplied migrations plan length: {len(plan)}")
        for migration, backwards in plan:
            res.append(f" - Unapplied: {migration.app}.{migration.name}")
            
        if plan or request.GET.get('migrate') == 'true':
            res.append("\n>>> Running migrate command...")
            import io
            buf = io.StringIO()
            call_command('migrate', no_input=True, stdout=buf, stderr=buf)
            res.append("Migration Output:")
            res.append(buf.getvalue())
            
            tables_after = connection.introspection.table_names()
            res.append(f"\nTables in database after migrate ({len(tables_after)}):")
            for t in tables_after:
                res.append(f" - {t}")
                
            executor_after = MigrationExecutor(connection)
            executor_after.loader.build_graph()
            plan_after = executor_after.migration_plan(executor_after.loader.graph.leaf_nodes())
            res.append(f"Unapplied migrations plan length after: {len(plan_after)}")
            
            res.append("\n>>> Re-running create_division_users...")
            buf_users = io.StringIO()
            call_command('create_division_users', stdout=buf_users, stderr=buf_users)
            res.append("create_division_users Output:")
            res.append(buf_users.getvalue())
            
            res.append("\n>>> Re-running fix_user_passwords...")
            buf_pwd = io.StringIO()
            call_command('fix_user_passwords', stdout=buf_pwd, stderr=buf_pwd)
            res.append("fix_user_passwords Output:")
            res.append(buf_pwd.getvalue())
            
    except Exception as e:
        res.append("Error occurred:")
        res.append(traceback.format_exc())
        
    return HttpResponse("<pre>" + "\n".join(res) + "</pre>", content_type="text/html")

